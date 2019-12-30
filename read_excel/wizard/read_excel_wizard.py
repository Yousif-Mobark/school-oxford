# -*- coding: utf-8 -*-
from odoo import models, fields, api
import base64
import xlrd
from openpyxl import load_workbook
import pandas as pd
from io import BytesIO
from datetime import datetime

class ReadExcel(models.TransientModel):
    _name = 'read.excel'

    excel_file = fields.Binary('Excel File', required=True)


    @api.multi
    def import_excel_file(self):

        df = pd.read_excel( BytesIO(base64.decodestring(self.excel_file)),engine=None)
        # convert all the NAN values fo 0
        # engine = sqlalchemy.create_engine("postgresql://user@localhost:8069/dbname")
        # df.to_sql(engine, 'classification', if_exists='append', index=False)
        # print ("@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")
        df = df.fillna(0)
        # print (df)
        employess_id = df[  self.env.user.company_id.device_id_col_name  ]

        checkin_time = df[ self.env.user.company_id.checkIn_time_col_name ]
        checkin_date = df[ self.env.user.company_id.checkIn_date_col_name ] 

        CheckOut_time = df[ self.env.user.company_id.checkIn_time_col_name ]
        CheckOut_date = df[ self.env.user.company_id.checkOut_date_col_name ]

        attendance_obj = self.env['hr.attendance'] 
        # list_t = []
        for i in range(0, len(checkin_time) ):

            # list_t.append(checkin_date[i]+" "+ str(checkin_time[i]).split()[0])
            if self.give_me_my_id(employess_id[i]):
                attendance_obj.create({
                        'employee_id': str( self.give_me_my_id(employess_id[i])), #i tried with actual employee id in firt column
                        'check_in':  checkin_date[i]+" "+ str(checkin_time[i]).split()[0] if checkin_time[i] else False,
                        'check_out':   CheckOut_date[i]+" "+ str(CheckOut_time[i]).split()[0] if CheckOut_time[i] else False
                })


    def give_me_my_id(self, barcode):
        return self.env["hr.employee"].search([("barcode",'=',barcode)]).id

        # print(list_t)

        # # spreadsheet = workbook.active
        # # print ("@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")
        # # wb = xlrd.open_workbook(file_contents=base64.decodestring(self.excel_file))
        # df = pd.read_excel(base64.decodestring(self.excel_file))
        # # df.head()
        # print ("@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")
        # # sheet_names = wb.sheet_by_index(0)
        # print(df.head() ) # 10> CheckOut 
        # print ("@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")
        # # for sheet in wb.sheets():
        #     # print ("@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")
        #     # print (sheet)
        #     # print ("@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")
        #     # for row in range(sheet.nrows):
        #     #     res = self.env['hr.attendance'].create({
        #     #             'employee_id': sheet.cell(row,0).value, #i tried with actual employee id in firt column
        #     #             'check_in': sheet.cell(row,1).value,
        #     #             'check_out': sheet.cell(row,2).value
        #     #             })